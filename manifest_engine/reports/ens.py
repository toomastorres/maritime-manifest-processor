"""
ens.py - Escritor del ENS (debit note, EXPO). Monto fijo 15 USD por BL.
"""

from openpyxl import Workbook
from openpyxl.styles import Font

from .. import config
from .styles import (
    set_column_widths, BORDER_BOTTOM_THIN, BORDER_DOUBLE,
    FONT_CALIBRI_11, FONT_CALIBRI_14B, FONT_CALIBRI_16B, ALIGN_CENTER,
)


def _voyage_fmt(voyage):
    """0625 -> 06/25."""
    return f"{voyage[:2]}/{voyage[2:]}" if voyage and len(voyage) == 4 else (voyage or "")


def generar_ens(bls_expo, ship_name, voyage, output_path, cfg=None):
    wb = Workbook()
    ws = wb.active

    ens_bls = [bl for bl in bls_expo if bl.charge("ENS").monto > 0]
    loading_port = ens_bls[0].port_of_loading if ens_bls else "Zarate"
    ws.title = loading_port[:31].title()

    set_column_widths(ws, {"A": 6, "B": 21, "C": 34, "D": 21, "E": 10, "F": 18})

    ws["B1"] = "Av. del Puerto 1000 - Buenos Aires (Argentina)"
    ws["B1"].font = FONT_CALIBRI_11
    ws["B3"] = "Messrs"
    ws["B3"].font = Font(name="Calibri", size=12, bold=True, italic=True)
    ws["C3"] = "Atlantic Maritime Agency"
    ws["C3"].font = Font(name="Calibri", size=14, bold=True)
    ws["C3"].border = BORDER_BOTTOM_THIN

    # Caja DEBIT NOTE (doble borde)
    ws.merge_cells("B13:E15")
    ws["B13"] = "DEBIT NOTE"
    ws["B13"].font = FONT_CALIBRI_16B
    ws["B13"].alignment = ALIGN_CENTER
    ws.merge_cells("F13:F15")
    ws["F13"] = "u$s"
    ws["F13"].font = FONT_CALIBRI_16B
    ws["F13"].alignment = ALIGN_CENTER
    for r in range(13, 16):
        for c in range(2, 7):
            ws.cell(row=r, column=c).border = BORDER_DOUBLE
    for r in (1, 3, 13, 14, 15):
        ws.row_dimensions[r].height = 18

    ws["B17"] = f"{ship_name} {_voyage_fmt(voyage)}".strip()
    ws["B17"].font = FONT_CALIBRI_14B
    ws["D17"] = "BL"
    ws["D17"].font = FONT_CALIBRI_14B

    row = 18
    for bl in ens_bls:
        ws.cell(row=row, column=3,
                value=f"{bl.port_of_loading} / {bl.port_of_discharge}").font = FONT_CALIBRI_11
        ws.cell(row=row, column=4, value=bl.bl_no.replace("[T]", "")).font = FONT_CALIBRI_14B
        ws.cell(row=row, column=5, value="ENS").font = FONT_CALIBRI_11
        ws.cell(row=row, column=6, value=config.ENS_FIXED_AMOUNT).font = FONT_CALIBRI_14B
        ws.row_dimensions[row].height = 18.75
        row += 1

    row += 1
    ws.cell(row=row, column=5, value="TOTAL").font = FONT_CALIBRI_14B
    tot = ws.cell(row=row, column=6, value=f"=SUM(F18:F{row-2})")
    tot.font = FONT_CALIBRI_16B
    tot.border = BORDER_DOUBLE

    wb.save(output_path)
