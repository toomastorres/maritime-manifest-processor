"""
chassis.py - Reporte de chasis/VIN (una fila por VIN).

Lista cada chasis extraído de los BL de autos/RoRo. Layout:
  A = B/L
  B = MARCA   (extract_make sobre la descripción; fallback al consignee/shipper)
  C = TIPO    (New/Used Car/Van/RoRo del BL)
  D = VIN/CHASIS

Solo incluye BL con vehículos que efectivamente imprimieron VIN. Los BL de autos
SIN VIN en el manifiesto NO ensucian este reporte: se reportan aparte en la hoja
REVISIÓN (ver validation.py). Una hoja IMPO y/o EXPO según el viaje.
"""

from openpyxl import Workbook

from .. import config
from ..domain.cargo import extract_make, detect_brand
from .styles import (
    apply_style_range, set_column_widths,
    BORDER_THIN, BORDER_MEDIUM, FILL_HEADER,
    FONT_ARIAL_10, FONT_ARIAL_10B, FONT_ARIAL_12B,
)


def _marca(bl):
    return extract_make(bl.description_lines) or detect_brand(bl.entity, bl.entity) or ""


def _tipo(bl):
    """Etiqueta del tipo de vehículo del BL (puede ser combinado)."""
    if not bl.cargo.vehicles:
        return ""
    return " + ".join(k.upper() for k in bl.cargo.vehicles)


def _write_sheet(wb, title, bls, op_type, ship, voyage):
    ws = wb.create_sheet(title=title[:31])
    set_column_widths(ws, {"A": 16, "B": 22, "C": 22, "D": 22})

    ws.cell(row=1, column=1, value=f"{ship} {voyage} - {op_type} - CHASIS/VIN").font = FONT_ARIAL_12B

    headers = {1: "B/L", 2: "MARCA", 3: "TIPO", 4: "VIN/CHASIS"}
    for col, val in headers.items():
        ws.cell(row=3, column=col, value=val)
    apply_style_range(ws, 3, 1, 4, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)

    row = 4
    for bl in bls:
        if not bl.cargo.vehicles or not bl.vins:
            continue  # sin vehículos o sin VIN impreso -> no va al detalle
        marca = _marca(bl)
        tipo = _tipo(bl)
        bl_disp = bl.bl_no.replace("[T]", "")
        for vin in bl.vins:
            ws.cell(row=row, column=1, value=bl_disp)
            ws.cell(row=row, column=2, value=marca)
            ws.cell(row=row, column=3, value=tipo)
            ws.cell(row=row, column=4, value=vin)
            apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_10, border=BORDER_THIN)
            row += 1

    if row == 4:  # sin ningún VIN en esta operación
        ws.cell(row=4, column=1, value="(sin VIN en el manifiesto)").font = FONT_ARIAL_10


def generar_chassis(impo_bls, expo_bls, ship, voyage, output_path, cfg=None):
    """Genera el reporte de chasis/VIN del viaje (hoja IMPO y/o EXPO)."""
    cfg = cfg or config.EngineConfig()
    wb = Workbook()
    wb.remove(wb.active)

    if impo_bls:
        _write_sheet(wb, "IMPO", impo_bls, "IMPO", ship, voyage)
    if expo_bls:
        _write_sheet(wb, "EXPO", expo_bls, "EXPO", ship, voyage)

    if not wb.sheetnames:
        wb.create_sheet(title="VACIO")

    wb.save(output_path)
