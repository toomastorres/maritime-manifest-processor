"""
comisiones.py - Escritor de comisiones, COMPARTIDO IMPO/EXPO.

Genera MontocomZte (IMPO) y MontoExpoZte (EXPO): secciones ROLLING / GENERAL
CARGO / CONTAINERS agrupadas por puerto, con USD/EURO, rates y ROE en Excel.

Los montos por sección salen de commissionable_split: un BL mixto aporta su parte
de contenedor a CONTAINERS y su parte general/autos a ROLLING o GENERAL CARGO.
"""

from collections import defaultdict

from openpyxl import Workbook

from .. import config
from ..rules.commission import commissionable_split, general_subcategory
from .styles import (
    apply_style_range, set_column_widths, write_money as _money,
    BORDER_THIN, BORDER_MEDIUM, BORDER_DOUBLE,
    FILL_HEADER, FILL_SECTION, FILL_TOTAL,
    FONT_ARIAL_10, FONT_ARIAL_10B, FONT_ARIAL_12B, FONT_ARIAL_14BI, FONT_ARIAL_16B,
    ALIGN_CENTER,
)


def generar_comisiones(bls, op_type, ship_name, voyage, output_path, cfg=None):
    cfg = cfg or config.EngineConfig()
    roe = cfg.roe
    rates = cfg.rates_for(op_type)
    is_impo = op_type.upper() == "IMPO"

    wb = Workbook()
    ws = wb.active
    ws.title = "COMISIONES"
    set_column_widths(ws, {"A": 18, "B": 18, "C": 20, "D": 14, "E": 14,
                           "F": 9, "G": 9, "H": 14, "I": 16, "J": 14})

    rolling = defaultdict(list)   # port -> [(bl, monto, moneda)]
    gc = defaultdict(list)
    cont = defaultdict(list)

    for bl in bls:
        port = bl.port_of_loading if is_impo else bl.port_of_discharge
        if port == "Nulo":
            port = "UNKNOWN"
        split = commissionable_split(bl)
        if bl.cargo.containers:
            cont[port].append((bl, split["container"], split["container_moneda"]))
        # parte general/autos
        if split["general"] > 0 or not bl.cargo.containers:
            amt = split["general"]
            moneda = split["general_moneda"]
            if general_subcategory(bl) == "cars":
                rolling[port].append((bl, amt, moneda))
            else:
                gc[port].append((bl, amt, moneda))

    row = 2
    rolling_cell = gc_cell = cont_cell = None

    # ─── ROLLING ───
    if rolling:
        row = _section_header(ws, row, "ROLLING COMMISSION")
        total_rows = []
        for port, items in rolling.items():
            ws.cell(row=row, column=1, value=port).font = FONT_ARIAL_10B
            apply_style_range(ws, row, 1, 4, fill=FILL_HEADER, border=BORDER_THIN)
            row += 2
            _mini_head(ws, row, ["B/L", "CARS", "USD", "EURO"])
            row += 1
            ws.cell(row=row, column=2, value="UNITS")
            ws.cell(row=row, column=3, value=rates.rolling)
            ws.cell(row=row, column=4, value=rates.rolling)
            row += 1
            ws.cell(row=row, column=1, value="ROLLING").font = FONT_ARIAL_10B
            row += 1
            start = row
            for bl, amt, moneda in items:
                ws.cell(row=row, column=1, value=bl.bl_no.replace("[T]", ""))
                ws.cell(row=row, column=2, value=bl.cargo.vehicle_count)
                if amt:
                    _money(ws, row, 3 if moneda == "USD" else 4, amt)
                apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_10, border=BORDER_THIN)
                row += 1
            row += 1  # blanco antes del TOTAL (como la verdad)
            ws.cell(row=row, column=1, value="TOTAL ROLLING")
            for c in (2, 3, 4):
                ws.cell(row=row, column=c, value=f"=SUM({chr(64+c)}{start}:{chr(64+c)}{row-1})")
            apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_12B, border=BORDER_MEDIUM, fill=FILL_TOTAL)
            total_rows.append(row)
            row += 2
        rolling_cell, row = _grand_total(ws, row, total_rows, rates.rolling, roe, "ROLLING")

    # ─── GENERAL CARGO ───
    if gc:
        row = _section_header(ws, row, "GENERAL CARGO COMMISSION")
        total_rows = []
        for port, items in gc.items():
            ws.cell(row=row, column=1, value=port).font = FONT_ARIAL_10B
            apply_style_range(ws, row, 1, 4, fill=FILL_HEADER, border=BORDER_THIN)
            row += 2
            _mini_head(ws, row, ["B/L", "USD", "EURO", "RORO"])
            row += 1
            ws.cell(row=row, column=2, value=rates.general_cargo)
            ws.cell(row=row, column=3, value=rates.general_cargo)
            ws.cell(row=row, column=4, value="UNITS")
            row += 1
            ws.cell(row=row, column=1, value="GENERAL CARGO").font = FONT_ARIAL_10B
            row += 1
            start = row
            for bl, amt, moneda in items:
                ws.cell(row=row, column=1, value=bl.bl_no.replace("[T]", ""))
                if amt:
                    _money(ws, row, 2 if moneda == "USD" else 3, amt)
                if bl.cargo.roro_count:
                    ws.cell(row=row, column=4, value=bl.cargo.roro_count)
                apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_10, border=BORDER_THIN)
                row += 1
            row += 1  # blanco antes del TOTAL
            ws.cell(row=row, column=1, value="TOTAL")
            for c in (2, 3, 4):
                ws.cell(row=row, column=c, value=f"=SUM({chr(64+c)}{start}:{chr(64+c)}{row-1})")
            apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_12B, border=BORDER_MEDIUM, fill=FILL_TOTAL)
            total_rows.append(row)
            row += 2
        gc_cell, row = _grand_total_gc(ws, row, total_rows, rates.general_cargo, roe)

    # ─── CONTAINERS ───
    if cont:
        row = _section_header(ws, row, "CONTAINERS COMMISSION", width=8)
        total_rows = []
        for port, items in cont.items():
            ws.cell(row=row, column=1, value=port).font = FONT_ARIAL_10B
            apply_style_range(ws, row, 1, 8, fill=FILL_HEADER, border=BORDER_THIN)
            row += 2
            _mini_head(ws, row, ["B/L", "USD", "EURO", "CTRS", "CTRS", "COMM", "MINIMUM", "FINAL"])
            row += 1
            ws.cell(row=row, column=2, value=rates.containers)
            ws.cell(row=row, column=3, value=rates.containers)
            ws.cell(row=row, column=7, value=f"USD {int(rates.container_min_per_unit)} X EACH")
            rates_row = row
            row += 1
            ws.cell(row=row, column=1, value="CONTAINER - MAFIS").font = FONT_ARIAL_10B
            row += 1
            start = row
            for bl, amt, moneda in items:
                ws.cell(row=row, column=1, value=bl.bl_no.replace("[T]", ""))
                n = bl.cargo.container_count
                if moneda == "USD":
                    _money(ws, row, 2, amt if amt else 0.0)
                    ws.cell(row=row, column=4, value=n)
                    ws.cell(row=row, column=6, value=f"=B{row}*B{rates_row}")
                else:
                    _money(ws, row, 3, amt if amt else 0.0)
                    ws.cell(row=row, column=5, value=n)
                    ws.cell(row=row, column=6, value=f"=C{row}*C{rates_row}*{roe}")
                col_ctr = "D" if moneda == "USD" else "E"
                ws.cell(row=row, column=7, value=f"={col_ctr}{row}*{int(rates.container_min_per_unit)}")
                ws.cell(row=row, column=8, value=f"=MAX(G{row},H{row})")
                apply_style_range(ws, row, 1, 8, font=FONT_ARIAL_10, border=BORDER_THIN)
                row += 1
            row += 1  # blanco antes del TOTAL
            ws.cell(row=row, column=1, value="TOTAL CONT / MAFIS")
            for c in (2, 3, 4, 5):
                ws.cell(row=row, column=c, value=f"=SUM({chr(64+c)}{start}:{chr(64+c)}{row-1})")
            ws.cell(row=row, column=8, value=f"=SUM(H{start}:H{row-1})")
            apply_style_range(ws, row, 1, 8, font=FONT_ARIAL_12B, border=BORDER_MEDIUM, fill=FILL_TOTAL)
            total_rows.append(row)
            row += 2
        # gran total contenedores
        if total_rows:
            refs = "+".join(f"H{r}" for r in total_rows)
            ws.cell(row=row, column=1, value="TOTAL COMMISSION CONTAINERS").font = FONT_ARIAL_12B
            ws.cell(row=row, column=2, value=f"={refs}")
            ws.cell(row=row, column=2).font = FONT_ARIAL_16B
            cont_cell = f"B{row}"
            row += 2

    # ─── GRAND TOTAL ───
    ws.cell(row=row, column=1, value="GRAND TOTAL:").font = FONT_ARIAL_16B
    parts = [c for c in (rolling_cell, gc_cell, cont_cell) if c]
    if parts:
        ws.cell(row=row, column=3, value="=" + "+".join(parts))
    ws.cell(row=row, column=3).font = FONT_ARIAL_16B
    ws.cell(row=row, column=3).border = BORDER_DOUBLE
    apply_style_range(ws, row, 1, 3, fill=FILL_TOTAL)

    from ..validation import revisar_bls
    from .revision import add_revision_sheet
    add_revision_sheet(wb, revisar_bls(bls, op_type, cfg))

    wb.save(output_path)


# ── helpers ──

def _section_header(ws, row, text, width=4):
    ws.cell(row=row, column=1, value=text).font = FONT_ARIAL_14BI
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    apply_style_range(ws, row, 1, width, border=BORDER_MEDIUM, fill=FILL_SECTION)
    ws.row_dimensions[row].height = 28
    return row + 2


def _mini_head(ws, row, labels):
    ws.cell(row=row - 1, column=3, value="AMOUNTS")
    ws.cell(row=row - 1, column=4, value="AMOUNTS")
    for i, lab in enumerate(labels):
        ws.cell(row=row, column=1 + i, value=lab)
    apply_style_range(ws, row, 1, len(labels), font=FONT_ARIAL_10B, border=BORDER_THIN, alignment=ALIGN_CENTER)


def _grand_total(ws, row, total_rows, rate, roe, label):
    if not total_rows:
        return None, row
    c_refs = "+".join(f"C{r}" for r in total_rows)
    d_refs = "+".join(f"D{r}" for r in total_rows)
    ws.cell(row=row, column=1, value="TOTAL")
    ws.cell(row=row, column=3, value=f"=({c_refs})*{rate}")
    ws.cell(row=row, column=4, value=f"=({d_refs})*{rate}*{roe}")
    comm_row = row
    row += 1
    ws.cell(row=row, column=1, value=f"TOTAL COMMISSION {label}").font = FONT_ARIAL_12B
    ws.cell(row=row, column=3, value=f"=C{comm_row}+D{comm_row}")
    ws.cell(row=row, column=3).font = FONT_ARIAL_16B
    cell = f"C{row}"
    row += 2
    ws.cell(row=row, column=1, value="ROE: EU 1,00 =")
    ws.cell(row=row, column=2, value=roe)
    row += 2
    return cell, row


def _grand_total_gc(ws, row, total_rows, rate, roe):
    if not total_rows:
        return None, row
    b_refs = "+".join(f"B{r}" for r in total_rows)
    c_refs = "+".join(f"C{r}" for r in total_rows)
    ws.cell(row=row, column=1, value="COMMISSION")
    ws.cell(row=row, column=2, value=f"=({b_refs})*{rate}")
    ws.cell(row=row, column=3, value=f"=({c_refs})*{rate}*{roe}")
    comm_row = row
    row += 1
    ws.cell(row=row, column=1, value="TOTAL COMMISSION GENERAL CARGO").font = FONT_ARIAL_12B
    ws.cell(row=row, column=2, value=f"=B{comm_row}+C{comm_row}")
    ws.cell(row=row, column=2).font = FONT_ARIAL_16B
    cell = f"B{row}"
    row += 2
    ws.cell(row=row, column=1, value="ROE: EU 1,00 =")
    ws.cell(row=row, column=2, value=roe)
    row += 2
    return cell, row
