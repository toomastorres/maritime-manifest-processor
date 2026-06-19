"""
por_marca.py - Reporte resumen de autos por marca (por viaje).

Agrega los BL de vehículos por marca. Layout:
  A = MARCA
  B = CANT. AUTOS   (declarado: suma de vehículos del manifiesto, cantidad real)
  C = CANT. VIN     (chasis efectivamente capturados; <= autos si el manifiesto
                     no imprimió todos los VIN)
  D = CANT. BL      (cuántos BL de esa marca)

El detalle de cada chasis por marca está en el reporte Chasis (columna MARCA).
Una hoja IMPO y/o EXPO según el viaje. Fila TOTAL al final.
"""

from collections import defaultdict

from openpyxl import Workbook

from .. import config
from ..domain.cargo import extract_make, detect_brand
from .styles import (
    apply_style_range, set_column_widths,
    BORDER_THIN, BORDER_MEDIUM, FILL_HEADER,
    FONT_ARIAL_10, FONT_ARIAL_10B, FONT_ARIAL_12B,
)

SIN_MARCA = "(sin marca)"


def _marca(bl):
    return extract_make(bl.description_lines) or detect_brand(bl.entity, bl.entity) or SIN_MARCA


def _write_sheet(wb, title, bls, op_type, ship, voyage):
    ws = wb.create_sheet(title=title[:31])
    set_column_widths(ws, {"A": 26, "B": 16, "C": 14, "D": 12})

    ws.cell(row=1, column=1, value=f"{ship} {voyage} - {op_type} - AUTOS POR MARCA").font = FONT_ARIAL_12B

    headers = {1: "MARCA", 2: "CANT. AUTOS", 3: "CANT. VIN", 4: "CANT. BL"}
    for col, val in headers.items():
        ws.cell(row=3, column=col, value=val)
    apply_style_range(ws, 3, 1, 4, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)

    agg = defaultdict(lambda: {"autos": 0, "vins": 0, "bls": 0})
    for bl in bls:
        if not bl.cargo.vehicles:
            continue
        m = _marca(bl)
        agg[m]["autos"] += bl.cargo.vehicle_count
        agg[m]["vins"] += len(bl.vins)
        agg[m]["bls"] += 1

    row = 4
    # Orden por cantidad de autos desc.
    for marca, d in sorted(agg.items(), key=lambda kv: -kv[1]["autos"]):
        ws.cell(row=row, column=1, value=marca)
        ws.cell(row=row, column=2, value=d["autos"])
        ws.cell(row=row, column=3, value=d["vins"])
        ws.cell(row=row, column=4, value=d["bls"])
        apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_10, border=BORDER_THIN)
        row += 1

    if agg:
        ws.cell(row=row, column=1, value="TOTAL")
        ws.cell(row=row, column=2, value=f"=SUM(B4:B{row-1})")
        ws.cell(row=row, column=3, value=f"=SUM(C4:C{row-1})")
        ws.cell(row=row, column=4, value=f"=SUM(D4:D{row-1})")
        apply_style_range(ws, row, 1, 4, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)
    else:
        ws.cell(row=4, column=1, value="(sin autos en este manifiesto)").font = FONT_ARIAL_10


def generar_por_marca(impo_bls, expo_bls, ship, voyage, output_path, cfg=None):
    """Genera el resumen de autos por marca del viaje (hoja IMPO y/o EXPO)."""
    cfg = cfg or config.EngineConfig()
    wb = Workbook()
    wb.remove(wb.active)

    if impo_bls:
        _write_sheet(wb, "IMPO", impo_bls, "IMPO", ship, voyage)
    if expo_bls:
        _write_sheet(wb, "EXPO", expo_bls, "EXPO", ship, voyage)

    if not wb.sheetnames:
        wb.create_sheet(title="VACIO")

    wb.save(output_path)
