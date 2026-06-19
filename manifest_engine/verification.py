"""
verification.py - Harness de comparación celda-por-celda contra la verdad exacta.
"""

import openpyxl


def _norm(v):
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def _num(v):
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def _first_marker_row(ws, marker="PUERTO:"):
    for r in range(1, min(ws.max_row, 50) + 1):
        if _norm(ws.cell(r, 1).value) == marker:
            return r
    return None


def diff_sheet(ws_gen, ws_truth, tol=0.5, offset=0):
    """
    Devuelve lista de diferencias (fila_gen, col, gen, truth).
    `offset` desplaza verticalmente la verdad: truth(r+offset) vs gen(r).
    """
    diffs = []
    max_row = max(ws_gen.max_row, ws_truth.max_row)
    max_col = max(ws_gen.max_column, ws_truth.max_column)
    for r in range(1, max_row + 1):
        tr = r + offset
        if tr < 1:
            continue
        for c in range(1, max_col + 1):
            g = _norm(ws_gen.cell(r, c).value)
            t = _norm(ws_truth.cell(tr, c).value)
            if g == t:
                continue
            gn, tn = _num(g), _num(t)
            if gn is not None and tn is not None and abs(gn - tn) <= tol:
                continue
            diffs.append((r, c, g, t))
    return diffs


def auto_offset(ws_gen, ws_truth):
    """Offset vertical = fila del primer 'PUERTO:' en truth menos en gen."""
    g = _first_marker_row(ws_gen)
    t = _first_marker_row(ws_truth)
    if g is not None and t is not None:
        return t - g
    return 0


def diff_workbooks(gen_path, truth_path, tol=0.5, max_report=40):
    wb_g = openpyxl.load_workbook(gen_path)
    wb_t = openpyxl.load_workbook(truth_path)
    print(f"GEN sheets:   {wb_g.sheetnames}")
    print(f"TRUTH sheets: {wb_t.sheetnames}")
    total_cells = total_diffs = 0
    for name_t in wb_t.sheetnames:
        # emparejar por nombre (ignorando espacios al final)
        match = next((n for n in wb_g.sheetnames if n.strip() == name_t.strip()), None)
        if match is None:
            print(f"\n[SHEET FALTANTE EN GEN] {name_t!r}")
            continue
        ws_t = wb_t[name_t]
        ws_g = wb_g[match]
        diffs = diff_sheet(ws_g, ws_t, tol)
        cells = ws_t.max_row * ws_t.max_column
        total_cells += cells
        total_diffs += len(diffs)
        print(f"\n=== {name_t!r}: {len(diffs)} diffs / ~{cells} celdas ===")
        from openpyxl.utils import get_column_letter
        for r, c, g, t in diffs[:max_report]:
            print(f"   {get_column_letter(c)}{r}: gen={g!r}  truth={t!r}")
    pct = 100 * (1 - total_diffs / total_cells) if total_cells else 0
    print(f"\nTOTAL: {total_diffs} diffs / ~{total_cells} celdas  ({pct:.1f}% coincidencia)")
    return total_diffs
