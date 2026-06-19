"""
styles.py - Constantes de estilo Excel (migradas de manifest_processor_v3.py).
"""

from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

_thin = Side(style="thin")
_medium = Side(style="medium")
_double = Side(style="double")

BORDER_THIN = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_MEDIUM = Border(left=_medium, right=_medium, top=_medium, bottom=_medium)
BORDER_DOUBLE = Border(left=_double, right=_double, top=_double, bottom=_double)
BORDER_BOTTOM_THIN = Border(bottom=_thin)
BORDER_BOTTOM_MEDIUM = Border(bottom=_medium)
BORDER_BOTTOM_DOUBLE = Border(bottom=_double)

FONT_ARIAL_10 = Font(name="Arial", size=10)
FONT_ARIAL_10B = Font(name="Arial", size=10, bold=True)
FONT_ARIAL_12B = Font(name="Arial", size=12, bold=True)
FONT_ARIAL_13BI = Font(name="Arial", size=13, bold=True, italic=True)
FONT_ARIAL_14BI = Font(name="Arial", size=14, bold=True, italic=True)
FONT_ARIAL_16B = Font(name="Arial", size=16, bold=True)
FONT_CALIBRI_11 = Font(name="Calibri", size=11)
FONT_CALIBRI_14B = Font(name="Calibri", size=14, bold=True)
FONT_CALIBRI_16B = Font(name="Calibri", size=16, bold=True)

FILL_HEADER = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_SECTION = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_TOTAL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def apply_style_range(ws, row, col_start, col_end, font=None, border=None, fill=None, alignment=None):
    """Aplica estilo a un rango de celdas en una fila."""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font:
            cell.font = font
        if border:
            cell.border = border
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment


def set_column_widths(ws, widths):
    """Fija anchos de columna desde un dict {letra: ancho}."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


# Formato de número para montos de moneda. El separador se muestra según el
# locale de Excel; en es-AR un número 2790.0 con '#,##0.00' se ve "2.790,00".
MONEY_FMT = "#,##0.00"


def money_format(currency=None):
    """Formato de celda para un monto; incluye la moneda si no es USD."""
    if currency and currency != "USD":
        return f'#,##0.00" {currency}"'
    return MONEY_FMT


def write_money(ws, row, col, value, currency=None):
    """
    Escribe un monto como NÚMERO con formato de celda (no como texto), para que
    quede bien estructurado (sumable/calculable). Devuelve la celda.
    """
    cell = ws.cell(row=row, column=col, value=round(float(value), 2))
    cell.number_format = money_format(currency)
    return cell


def abreviar_puerto(puerto):
    """ZARATE -> ZTE, BUENOS AIRES -> BUE, etc."""
    from .. import config
    if not puerto or puerto == "Nulo":
        return ""
    p = puerto.upper().strip()
    if p in config.PORT_ABBREVIATIONS:
        return config.PORT_ABBREVIATIONS[p]
    return p[:3]
