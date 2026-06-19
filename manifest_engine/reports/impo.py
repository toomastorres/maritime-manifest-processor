"""
impo.py - Escritor de la planilla principal IMPO (hoja por puerto + hoja TOLL).

Portado de generar_planilla_impo() de manifest_processor_v3.py, operando sobre
BLRecord y tomando el divisor del PESO de la configuración (tarifa fija suelta).
"""

from collections import defaultdict

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from .. import config
from ..domain.cargo import format_cargo_output, cargo_category
from ..rules.biblia import format_val, calcular_totales_impo
from .styles import (
    apply_style_range, set_column_widths, abreviar_puerto, MONEY_FMT,
    BORDER_THIN, BORDER_MEDIUM, FILL_HEADER, FILL_YELLOW,
    FONT_ARIAL_10, FONT_ARIAL_10B, FONT_ARIAL_12B,
)

# Columnas de moneda en la planilla IMPO (F..O salvo E=DTN).
_MONEY_COLS_IMPO = (6, 7, 8, 9, 10, 11, 12, 13, 14, 15)


def _fmt_money_row(ws, row, cols=_MONEY_COLS_IMPO):
    """Aplica formato de número a las celdas de monto que sean numéricas/fórmula
    (deja el texto como 'Prepaid' o el nombre de puerto sin tocar)."""
    for c in cols:
        cell = ws.cell(row=row, column=c)
        v = cell.value
        if isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("=")):
            cell.number_format = MONEY_FMT


def _fmt(x):
    """Número como en la verdad exacta: entero sin decimales, si no hasta 2."""
    x = round(x, 2)
    s = f"{x:.2f}"
    if s.endswith(".00"):
        return s[:-3]
    if s.endswith("0"):
        return s[:-1]
    return s


def _comm_formula(gross, disc):
    """
    Celda del Monto Comisión: '0,00' si neto<=0 (clamp, nunca negativo);
    fórmula =gross-disc si hay descuento; número si no.
    """
    g = round(gross, 2)
    d = round(disc, 2)
    if round(g - d, 2) <= 0.005:
        return 0.0
    if d > 1e-9:
        return f"={_fmt(g)}-{_fmt(d)}"
    return g


def _write_totals_impo(ws, row, bl):
    """Escribe DOLARES/EURO PREPAID/COLLECT + BAF + COMM con fórmulas-componentes."""
    basic = bl.charge("Basic FRT")
    ot_sum = sum(cl.total for cl in bl.charge_lines
                 if cl.total > 0 and any(cl.desc.startswith(p) for p in config.COMMISSIONABLE_EXTRA))
    gross = basic.monto + ot_sum
    disc = -bl.total_ajustes_negativos  # positivo
    comm_cell = _comm_formula(gross, disc)

    baf = bl.charge("BAF")
    baf_collect = baf.letra == "C" and baf.monto
    total_ba_usd = bl.totals.get("Total Buenos Aires USD", 0) or 0
    total_ba_eur = bl.totals.get("Total Buenos Aires EUR", 0) or 0

    if basic.letra == "P":
        if basic.moneda == "USD":
            ws.cell(row=row, column=10, value=comm_cell)   # J DOLARES PREPAID
        elif basic.moneda == "EUR":
            ws.cell(row=row, column=12, value=comm_cell)   # L EURO PREPAID
    elif basic.letra == "C":
        ws.cell(row=row, column=15, value=comm_cell)        # O COMM (por fuera)
        if baf_collect:
            ws.cell(row=row, column=14, value=round(baf.monto, 2))  # N BAF discriminado
        if basic.moneda == "USD":
            # K COLLECT = Total BA USD - BAF(N) - THC(C) - S/C(C)
            extra = 0.0
            if bl.charge("THC").letra == "C":
                extra += bl.charge("THC").monto
            if bl.charge("Sweeping").letra == "C":
                extra += bl.charge("Sweeping").monto
            if bl.charge("Toll").letra == "C":
                extra += bl.charge("Toll").monto
            f = f"={_fmt(total_ba_usd)}"
            if baf_collect:
                f += f"-N{row}"
            if abs(extra) > 1e-9:
                f += f"-{_fmt(extra)}"
            ws.cell(row=row, column=11, value=f if (baf_collect or abs(extra) > 1e-9) else round(total_ba_usd, 2))
        elif basic.moneda == "EUR":
            # M COLLECT = Total BA EUR - BAF(N)
            if baf_collect:
                ws.cell(row=row, column=13, value=f"={_fmt(total_ba_eur)}-N{row}")
            else:
                ws.cell(row=row, column=13, value=round(total_ba_eur, 2))


def generar_planilla_impo(bls, ship_name, voyage, output_path, cfg=None):
    cfg = cfg or config.EngineConfig()
    tonnage_factor = cfg.tonnage_factor

    wb = Workbook()
    wb.remove(wb.active)

    by_port = defaultdict(list)
    for bl in bls:
        by_port[bl.port_of_discharge].append(bl)

    toll_refs = {}  # {sheet: {cat: [rows]}}

    for port_name, port_bls in by_port.items():
        ws = wb.create_sheet(title=port_name[:31])
        toll_refs[ws.title] = {"containers": [], "projects": [], "cars": []}
        dtn = abreviar_puerto(port_name)

        set_column_widths(ws, {"A": 14, "B": 1, "C": 28, "D": 10, "E": 6, "F": 12,
                               "G": 10, "H": 10, "I": 10, "J": 12, "K": 12,
                               "L": 12, "M": 12, "N": 10, "O": 12})

        by_loading = defaultdict(list)
        for bl in port_bls:
            by_loading[bl.port_of_loading].append(bl)

        row = 1
        # La verdad exacta ordena los puertos de carga alfabéticamente.
        for loading_port, lp_bls in sorted(by_loading.items()):
            total_20 = sum(bl.thc_20.monto / cfg.tariffs.thc_20
                           if bl.thc_20.monto and bl.thc_20.letra else 0 for bl in lp_bls)
            total_40 = sum(bl.thc_40.monto / cfg.tariffs.thc_40
                           if bl.thc_40.monto and bl.thc_40.letra else 0 for bl in lp_bls)

            row += 1
            ws.cell(row=row, column=1, value="PUERTO:").font = FONT_ARIAL_10B
            ws.cell(row=row, column=3, value=loading_port).font = FONT_ARIAL_10B
            ws.cell(row=row, column=8, value=f"{dtn} x 20'").font = FONT_ARIAL_10B
            ws.cell(row=row, column=9, value=int(total_20) if total_20 else 0).font = FONT_ARIAL_10B
            ws.cell(row=row, column=10, value=f"{dtn} x 40'").font = FONT_ARIAL_10B
            ws.cell(row=row, column=11, value=int(total_40) if total_40 else 0).font = FONT_ARIAL_10B
            apply_style_range(ws, row, 1, 15, border=BORDER_MEDIUM, fill=FILL_HEADER)

            row += 1
            headers_r3 = {1: "B/L", 3: "CARGA", 4: "PESO", 5: "DTN", 6: "TOLL", 7: "THC", 9: "SWP",
                          10: "DOLARES", 12: "EURO"}
            for col, val in headers_r3.items():
                ws.cell(row=row, column=col, value=val)
            apply_style_range(ws, row, 1, 15, font=FONT_ARIAL_10B, border=BORDER_THIN)

            row += 1
            headers_r4 = {7: "20'", 8: "40'", 10: "PREPAID", 11: "COLLECT", 12: "PREPAID",
                          13: "COLLECT", 14: "BAF", 15: "COMM"}
            for col, val in headers_r4.items():
                ws.cell(row=row, column=col, value=val)
            apply_style_range(ws, row, 1, 15, font=FONT_ARIAL_10B, border=BORDER_THIN)

            row += 1
            data_start = row
            prev_bl_number = None
            for bl in lp_bls:
                cargo_text = format_cargo_output(bl.cargo, bl.entity, bl.description_lines)
                toll_val = bl.charge("Toll").monto
                toll_letra = bl.charge("Toll").letra
                has_containers = bool(bl.cargo.containers)
                n_cont = bl.cargo.container_count

                bl_display = bl.bl_no
                if bl_display == prev_bl_number:
                    bl_display = '""'

                ws.cell(row=row, column=1, value=bl_display)
                ws.cell(row=row, column=3, value=n_cont if has_containers else cargo_text)

                # PESO
                if has_containers:
                    ws.cell(row=row, column=4, value=n_cont)
                else:
                    if toll_val:
                        ws.cell(row=row, column=4, value=f"=F{row}/{tonnage_factor}")
                    elif bl.weight > 0:
                        ws.cell(row=row, column=4, value=round(bl.weight / 1000, 3))

                ws.cell(row=row, column=5, value=dtn)

                # TOLL: con monto -> número; prepaid -> puerto de carga; sin
                # dato -> celda VACÍA (no 0).
                if toll_val and toll_letra != "P":
                    ws.cell(row=row, column=6, value=toll_val)
                elif toll_val and toll_letra == "P":
                    ws.cell(row=row, column=6, value=bl.port_of_loading)

                # THC 20'/40' y SWP. Sin dato -> celda VACÍA (no 0); 'Prepaid'
                # queda como texto cuando corresponde.
                thc20_cell = format_val(bl.thc_20.monto, bl.thc_20.letra, item_type="THC")
                thc40_cell = format_val(bl.thc_40.monto, bl.thc_40.letra, item_type="THC")
                swp = bl.charge("Sweeping")
                swp_cell = format_val(swp.monto, swp.letra, item_type="SWP")
                if thc20_cell != "":
                    ws.cell(row=row, column=7, value=thc20_cell)
                if thc40_cell != "":
                    ws.cell(row=row, column=8, value=thc40_cell)
                if swp_cell != "":
                    ws.cell(row=row, column=9, value=swp_cell)

                # Totales (fórmulas con componentes, como la verdad exacta)
                _write_totals_impo(ws, row, bl)

                apply_style_range(ws, row, 1, 15, font=FONT_ARIAL_10, border=BORDER_THIN)
                _fmt_money_row(ws, row)

                # Sólo se contabilizan en la hoja TOLL las filas con TOLL
                # numérico > 0 (las prepaid/0 se excluyen; ver +887.5 manual).
                cat = cargo_category(bl.cargo)
                if cat and toll_val and toll_letra != "P":
                    toll_refs[ws.title][cat].append(row)

                prev_bl_number = bl.bl_no
                row += 1

            # Fila TOTALES del grupo: 1 blanco, TOTALES con =SUM(F..N), 1 blanco.
            row += 1
            ws.cell(row=row, column=3, value="TOTALES").font = FONT_ARIAL_10B
            for col in range(6, 15):  # F..N
                letter = get_column_letter(col)
                ws.cell(row=row, column=col,
                        value=f"=SUM({letter}{data_start}:{letter}{row - 1})")
            apply_style_range(ws, row, 1, 15, font=FONT_ARIAL_10B, border=BORDER_THIN)
            row += 1  # blanco entre TOTALES y el próximo grupo

    _toll_sheet(wb, toll_refs)

    # Hoja REVISIÓN con las alertas del viaje
    from ..validation import revisar_bls
    from .revision import add_revision_sheet
    add_revision_sheet(wb, revisar_bls(bls, "IMPO", cfg))

    wb.save(output_path)


def _toll_sheet(wb, toll_refs):
    ws_toll = wb.create_sheet(title="TOLL")
    set_column_widths(ws_toll, {"B": 18, "C": 14, "D": 14, "E": 14, "F": 14})

    ws_toll["B3"] = "///////////////////"
    ws_toll["C3"] = "CONTAINERS"
    ws_toll["D3"] = "PROJECTS"
    ws_toll["E3"] = "CARS"
    ws_toll["B4"] = "TOLL"
    apply_style_range(ws_toll, 3, 2, 5, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)
    apply_style_range(ws_toll, 4, 2, 5, font=FONT_ARIAL_10, border=BORDER_THIN)

    for col_idx, cat in [(3, "containers"), (4, "projects"), (5, "cars")]:
        formulas = []
        for sheet_name, refs in toll_refs.items():
            for r in refs[cat]:
                formulas.append(f"'{sheet_name}'!F{r}")
        if formulas:
            ws_toll.cell(row=4, column=col_idx, value="=" + "+".join(formulas))

    ws_toll["C7"] = "TOTAL:"
    ws_toll["D7"] = "=C4+D4+E4"
    ws_toll.cell(row=7, column=3).font = FONT_ARIAL_12B
    ws_toll.cell(row=7, column=4).font = FONT_ARIAL_12B
    ws_toll.cell(row=7, column=4).fill = FILL_YELLOW
    ws_toll.cell(row=7, column=4).border = BORDER_MEDIUM

    ws_toll["B10"] = "///////////////////"
    ws_toll["C10"] = "CONTAINERS"
    ws_toll["D10"] = "PROJECTS"
    ws_toll["E10"] = "RORO"
    ws_toll["F10"] = "CARS"
    ws_toll["B11"] = "AMOUNT"
    ws_toll["C11"] = "=C4/150"
    ws_toll["D11"] = 0
    ws_toll["E11"] = 0
    ws_toll["F11"] = 0
    apply_style_range(ws_toll, 10, 2, 6, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)
    apply_style_range(ws_toll, 11, 2, 6, font=FONT_ARIAL_10, border=BORDER_THIN)
