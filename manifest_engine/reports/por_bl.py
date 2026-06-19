"""
por_bl.py - Reporte "por B/L": un listado plano de cada Bill of Lading con su
cliente, puertos y Total Buenos Aires.

Layout (un archivo por viaje, columnas básicas):
  A = B/L
  B = CLIENTE   (consignee en IMPO, shipper en EXPO; bl.entity)
  C = PUERTO CARGA      (bl.port_of_loading)
  D = PUERTO DESCARGA   (bl.port_of_discharge)
  E = TOTAL BUENOS AIRES (importe; con código de moneda si no es USD)

Una hoja IMPO y/o una hoja EXPO según los manifiestos disponibles del viaje.
No recalcula nada: lee los campos ya extraídos por el parser.
"""

import re

from openpyxl import Workbook

from .. import config


def _clean_entity(entity):
    """
    Limpia el nombre de cliente para el listado: corta desde 'CUIT' en adelante
    (el nombre a veces comparte línea con el CUIT en el manifiesto) y saca
    paréntesis/puntuación colgada. No altera el `entity` compartido del BL.
    """
    if not entity or entity == "Nulo":
        return ""
    e = re.split(r"\s*[\(,]?\s*C\.?U\.?I\.?T", entity, maxsplit=1, flags=re.IGNORECASE)[0]
    return e.rstrip(" ,(-").strip()
from .styles import (
    apply_style_range, set_column_widths, write_money,
    BORDER_THIN, BORDER_MEDIUM, FILL_HEADER,
    FONT_ARIAL_10, FONT_ARIAL_10B, FONT_ARIAL_12B,
)


def _total_ba(bl, op_type):
    """
    Devuelve (monto, moneda) del Total Buenos Aires del BL.
    IMPO: el manifiesto trae 'Total Buenos Aires USD' y/o 'EUR' (condición C).
    EXPO: trae 'Total Buenos Aires Monto' + 'Total Buenos Aires Moneda' (cond. P).
    """
    if op_type == "IMPO":
        usd = bl.totals.get("Total Buenos Aires USD", 0) or 0
        eur = bl.totals.get("Total Buenos Aires EUR", 0) or 0
        if usd:
            return usd, "USD"
        if eur:
            return eur, "EUR"
        return 0, "USD"
    monto = bl.totals.get("Total Buenos Aires Monto", 0) or 0
    moneda = bl.totals.get("Total Buenos Aires Moneda", "USD") or "USD"
    return monto, moneda


def _write_sheet(wb, title, bls, op_type, ship, voyage):
    ws = wb.create_sheet(title=title[:31])
    set_column_widths(ws, {"A": 16, "B": 40, "C": 24, "D": 24, "E": 20})

    ws.cell(row=1, column=1, value=f"{ship} {voyage} - {op_type}").font = FONT_ARIAL_12B

    headers = {1: "B/L", 2: "CLIENTE", 3: "PUERTO CARGA",
               4: "PUERTO DESCARGA", 5: "TOTAL BUENOS AIRES"}
    for col, val in headers.items():
        ws.cell(row=3, column=col, value=val)
    apply_style_range(ws, 3, 1, 5, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)

    row = 4
    for bl in bls:
        monto, moneda = _total_ba(bl, op_type)
        ws.cell(row=row, column=1, value=bl.bl_no.replace("[T]", ""))
        ws.cell(row=row, column=2, value=_clean_entity(bl.entity))
        ws.cell(row=row, column=3, value=bl.port_of_loading)
        ws.cell(row=row, column=4, value=bl.port_of_discharge)
        if monto:
            write_money(ws, row, 5, monto, moneda)
        apply_style_range(ws, row, 1, 5, font=FONT_ARIAL_10, border=BORDER_THIN)
        row += 1


def generar_por_bl(impo_bls, expo_bls, ship, voyage, output_path, cfg=None):
    """
    Genera el reporte por B/L del viaje. `impo_bls`/`expo_bls` pueden ser None si
    el manifiesto correspondiente no existe; se crea una hoja por cada uno.
    """
    cfg = cfg or config.EngineConfig()
    wb = Workbook()
    wb.remove(wb.active)

    if impo_bls:
        _write_sheet(wb, "IMPO", impo_bls, "IMPO", ship, voyage)
    if expo_bls:
        _write_sheet(wb, "EXPO", expo_bls, "EXPO", ship, voyage)

    if not wb.sheetnames:  # nada que escribir
        wb.create_sheet(title="VACIO")

    wb.save(output_path)
