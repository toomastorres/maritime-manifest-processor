"""
expo.py - Escritor de la planilla principal EXPO (hoja por puerto de carga + TOLL).

Layout: hoja = puerto de carga (ZÁRATE); grupos por DESTINO (puerto de descarga).
Columnas:
  A=DESTINO B=B/L C=CARGA D/E=THC 20'/40' F=SWP G=ENS H=DOLARES INCL PREPAID
  COLLECT en 4 columnas de moneda: I=USD J=EUR K=GBP L=BRL
  M=ABROAD N=MONTO COMISIÓN O=% P==N*O/100 Q=MÍNIMO(=C*30) R=FINAL(max)
  S=CTRS T=AUTO  V=TOLL

Convenciones: THC/SWP/TOLL sin dato quedan VACÍOS (no 0). Montos como número con
formato (no texto).
"""

from collections import defaultdict

from openpyxl import Workbook

from .. import config
from ..domain.cargo import format_cargo_output, cargo_category
from .styles import (
    apply_style_range, set_column_widths, write_money, MONEY_FMT,
    BORDER_THIN, BORDER_MEDIUM, FILL_HEADER, FILL_TOTAL, FILL_YELLOW,
    FONT_ARIAL_10, FONT_ARIAL_10B, FONT_ARIAL_12B, ALIGN_CENTER,
)

# Columnas COLLECT por moneda (las 4 posibles).
COLLECT_COLS = (("USD", 9), ("EUR", 10), ("GBP", 11), ("BRL", 12))


def _num(x):
    """Número como en la verdad: entero sin decimales, si no hasta 2."""
    x = round(x, 2)
    s = f"{x:.2f}"
    if s.endswith(".00"):
        return s[:-3]
    if s.endswith("0"):
        return s[:-1]
    return s


def _collect_by_currency(bl):
    """Suma los Total[Pais] collect del BL por moneda -> {USD: x, EUR: y, ...}."""
    by_cur = defaultdict(float)
    for it in bl.collect:
        by_cur[it.get("Moneda", "USD")] += it.get("Monto", 0)
    return by_cur


def generar_planilla_expo(bls, ship_name, voyage, output_path, cfg=None):
    cfg = cfg or config.EngineConfig()
    wb = Workbook()
    wb.remove(wb.active)

    by_loading = defaultdict(list)
    for bl in bls:
        by_loading[bl.port_of_loading].append(bl)

    toll_refs = {}

    for loading_port, lp_bls in by_loading.items():
        ws = wb.create_sheet(title=loading_port[:31])
        toll_refs[ws.title] = {"containers": [], "projects": [], "cars": []}

        set_column_widths(ws, {"A": 17, "B": 16, "C": 24, "D": 10, "E": 10, "F": 8, "G": 8,
                               "H": 13, "I": 11, "J": 11, "K": 11, "L": 11, "M": 12,
                               "N": 12, "O": 5, "P": 12, "Q": 12, "R": 12, "S": 7,
                               "T": 8, "U": 3, "V": 12})

        # Encabezados (3 filas)
        ws.cell(row=2, column=1, value="PUERTO:").font = FONT_ARIAL_10B
        ws.cell(row=2, column=3, value=loading_port).font = FONT_ARIAL_10B
        apply_style_range(ws, 2, 1, 22, border=BORDER_MEDIUM, fill=FILL_HEADER)

        h3 = {1: "DESTINO", 2: "B/L", 3: "CARGA", 4: "THC", 5: "THC", 6: "SWP", 7: "ENS",
              8: "DOLARES", 14: "MONTO", 15: "%", 16: "CALCULO", 17: "MINIMO ",
              18: "CALCULO", 19: "CTRS", 20: "AUTO"}
        h4 = {8: "INCL ENS", 9: "COLLECT", 17: "CONT. USD", 18: "COMISION", 20: "RORO"}
        h5 = {4: "x 20'", 5: "x 40'", 8: "PREPAID", 9: "USD", 10: "EUR", 11: "GBP",
              12: "BRL", 13: "ABROAD", 14: "COMISION", 16: "COMISION", 17: "30 X CONT",
              18: "FINAL", 22: "TOLL"}
        for col, v in h3.items():
            ws.cell(row=3, column=col, value=v)
        for col, v in h4.items():
            ws.cell(row=4, column=col, value=v)
        for col, v in h5.items():
            ws.cell(row=5, column=col, value=v)
        ws.merge_cells(start_row=4, start_column=9, end_row=4, end_column=12)  # COLLECT
        ws.cell(row=4, column=9).alignment = ALIGN_CENTER
        for r in (3, 4, 5):
            apply_style_range(ws, r, 1, 22, font=FONT_ARIAL_10B, border=BORDER_THIN, alignment=ALIGN_CENTER)

        by_dest = defaultdict(list)
        for bl in lp_bls:
            by_dest[bl.port_of_discharge].append(bl)

        row = 6
        for dest, dgroup in by_dest.items():
            first = True
            for bl in dgroup:
                row = _write_expo_bl(ws, row, bl, dest if first else "", cfg,
                                     toll_refs[ws.title])
                first = False

            # Fila de totales del grupo (S=CTRS, T=AUTO)
            n_ctr = sum(bl.cargo.container_count for bl in dgroup)
            n_veh = sum(bl.cargo.vehicle_count for bl in dgroup)
            if n_ctr:
                ws.cell(row=row, column=19, value=n_ctr)
            if n_veh:
                ws.cell(row=row, column=20, value=n_veh)
            row += 1

    _toll_sheet_expo(wb, toll_refs)

    from ..validation import revisar_bls
    from .revision import add_revision_sheet
    add_revision_sheet(wb, revisar_bls(bls, "EXPO", cfg))

    wb.save(output_path)


def _toll_split(bl):
    cont = gen = 0.0
    for cl in bl.charge_lines:
        if cl.desc.startswith("s/c"):
            if cl.is_container_basis:
                cont += cl.total
            else:
                gen += cl.total
    return cont, gen


def _write_expo_bl(ws, row, bl, dest, cfg, toll_refs):
    """Escribe 1 o 2 filas: contenedor y/o general (BL mixto se parte)."""
    from ..domain.models import CargoData
    from ..rules.commission import commissionable_split, general_subcategory

    split = commissionable_split(bl)
    cont_toll, gen_toll = _toll_split(bl)
    has_cont = bool(bl.cargo.containers)
    is_rolling = general_subcategory(bl) == "cars"
    has_other = split["general"] > 0.005

    cat = cargo_category(bl.cargo)
    segments = []
    if has_cont:
        segments.append(("cont", bl.cargo.container_count, split["container"],
                         split["container_moneda"], cont_toll, 5, "containers"))
    if has_other:
        other = CargoData(vehicles=dict(bl.cargo.vehicles), general=dict(bl.cargo.general))
        ctext = format_cargo_output(other, bl.entity, bl.description_lines)
        pct = 1 if is_rolling else 5
        segments.append(("other", ctext, split["general"], split["general_moneda"],
                         gen_toll, pct, "cars" if is_rolling else (cat or "projects")))
    if not segments:  # sin comisionable: 1 fila con la carga que haya
        ctext = bl.cargo.container_count if has_cont else \
            format_cargo_output(bl.cargo, bl.entity, bl.description_lines)
        segments.append(("cont" if has_cont else "other", ctext, 0, "USD",
                         bl.charge("Toll").monto, 1 if is_rolling else 5,
                         "containers" if has_cont else (cat or "projects")))

    # Datos compartidos (line costs / totales) van en la primera fila
    total_ba = bl.totals.get("Total Buenos Aires Monto", 0) or 0
    thc20, thc40, swp = bl.thc_20, bl.thc_40, bl.charge("Sweeping")
    sum_p = (thc20.monto if thc20.letra == "P" else 0) + (thc40.monto if thc40.letra == "P" else 0) \
        + (swp.monto if swp.letra == "P" else 0)
    is_matriz = (total_ba - sum_p) < 0 and sum_p > 0
    abroad = bl.totals.get("Total Matriz Monto", "")
    abroad = float(abroad) if abroad else ""

    def matriz(ch):
        if ch.monto <= 0:
            return ""
        return "MATRIZ" if (is_matriz and ch.letra == "P") else ch.monto

    for i, (kind, cval, comis, moneda, toll, pct, catref) in enumerate(segments):
        primary = (i == 0)
        ws.cell(row=row, column=1, value=dest if primary else "")
        ws.cell(row=row, column=2, value=bl.bl_no.replace("[T]", "") if primary else '""')
        ws.cell(row=row, column=3, value=cval)

        if primary:
            # THC/SWP: sin dato -> celda VACÍA (no 0). 'MATRIZ' es texto.
            t20, t40, sw = matriz(thc20), matriz(thc40), matriz(swp)
            if t20 != "":
                ws.cell(row=row, column=4, value=t20)
            if t40 != "":
                ws.cell(row=row, column=5, value=t40)
            if sw != "":
                ws.cell(row=row, column=6, value=sw)
            ens = bl.charge("ENS").monto
            ws.cell(row=row, column=7, value=ens if ens else "***")
            # H = PREPAID = Total BA - THC(P) - SWP(P)
            if is_matriz or total_ba == 0:
                if total_ba:
                    ws.cell(row=row, column=8, value=total_ba)
            else:
                refs = ""
                if thc20.monto > 0 and thc20.letra == "P":
                    refs += f"-D{row}"
                if thc40.monto > 0 and thc40.letra == "P":
                    refs += f"-E{row}"
                if swp.monto > 0 and swp.letra == "P":
                    refs += f"-F{row}"
                if refs:
                    ws.cell(row=row, column=8, value=f"={_num(total_ba)}{refs}")
                elif total_ba:
                    ws.cell(row=row, column=8, value=total_ba)
            # COLLECT: una columna por moneda (USD/EUR/GBP/BRL), como número.
            coll = _collect_by_currency(bl)
            for cur, cn in COLLECT_COLS:
                if coll.get(cur):
                    write_money(ws, row, cn, coll[cur])
            # M = ABROAD
            if abroad != "":
                ws.cell(row=row, column=13, value=abroad)

        # N = MONTO COMISIÓN (número con formato; EUR muestra sufijo).
        if comis:
            write_money(ws, row, 14, comis, moneda)
        elif kind == "cont":
            write_money(ws, row, 14, 0.0)
        ws.cell(row=row, column=15, value=pct)
        ws.cell(row=row, column=16, value=f"=N{row}*O{row}/100")
        if kind == "cont":
            ws.cell(row=row, column=17, value=f"=C{row}*30")
            m_val = (comis or 0) * pct / 100
            n_val = (cval or 0) * 30
            ws.cell(row=row, column=18, value=f"=Q{row}" if n_val >= m_val else f"=P{row}")
        else:
            ws.cell(row=row, column=18, value=f"=P{row}")
        # V = TOLL (vacío si no hay)
        if toll:
            ws.cell(row=row, column=22, value=toll)
            toll_refs[catref].append(row)

        apply_style_range(ws, row, 1, 22, font=FONT_ARIAL_10, border=BORDER_THIN)
        # Formato de moneda en celdas numéricas/fórmula (no toca texto/vacío).
        for c in (4, 5, 6, 8, 13, 16, 17, 18, 22):
            cell = ws.cell(row=row, column=c)
            if isinstance(cell.value, (int, float)) or \
                    (isinstance(cell.value, str) and cell.value.startswith("=")):
                cell.number_format = MONEY_FMT
        row += 1

    return row


def _toll_sheet_expo(wb, toll_refs):
    ws = wb.create_sheet(title="TOLL")
    set_column_widths(ws, {"B": 18, "C": 14, "D": 14, "E": 14, "F": 14})
    ws["B3"] = "///////////////////"
    ws["C3"] = "CONTAINERS"
    ws["D3"] = "PROJECTS"
    ws["E3"] = "CARS"
    ws["B4"] = "TOLL"
    apply_style_range(ws, 3, 2, 5, font=FONT_ARIAL_10B, border=BORDER_MEDIUM, fill=FILL_HEADER)
    apply_style_range(ws, 4, 2, 5, font=FONT_ARIAL_10, border=BORDER_THIN)
    for col_idx, cat in [(3, "containers"), (4, "projects"), (5, "cars")]:
        formulas = []
        for sheet_name, refs in toll_refs.items():
            for r in refs[cat]:
                formulas.append(f"'{sheet_name}'!V{r}")
        if formulas:
            ws.cell(row=4, column=col_idx, value="=" + "+".join(formulas))
    ws["C7"] = "TOTAL:"
    ws["D7"] = "=C4+D4+E4"
    ws.cell(row=7, column=4).fill = FILL_YELLOW
    ws.cell(row=7, column=4).border = BORDER_MEDIUM
